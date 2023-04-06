import cv2
import os
import time
import openpyxl
import face_recognition

def add_the_face():
    ######### if this directory doesn't exist, make them #######
    try:
        os.makedirs('images')
    except:
        pass
    ############### take name and roll no as input #################
    print("Enter student's details:")
    roll = input('Roll number: ')
    while roll is not None and not roll.isdigit():
        roll = input('Roll number: ')
    name = input('Name: ')
    email = input('Email: ')
    parent_name = input('Parent name: ')
    parent_email = input('Parent email: ')

    imageName = roll + '_' + name.replace(' ', '_')

    ######### add image in folder ##################
    if imageName + '.png' in os.listdir('images') or imageName + '.jpg' in os.listdir('images'):
        print('User already exists')
        input('\nPress any key to continue...')

    else:
        cap = cv2.VideoCapture(0)
        i = 0
        print()
        for i in range(5):
            print(f'Capturing starts in {5-i} seconds...')
            time.sleep(1)
        print('Taking photo...')
        if not cap.isOpened():
            print('Camera not found')
        ret, frame = cap.read()
        cv2.imshow("Taking your picture, be steady", frame)
        facesCurrentFrame = face_recognition.face_locations(frame)
        if len(facesCurrentFrame) > 1:
            print('More than one face found')
            input('Press any key to continue...')
        if cv2.imwrite('images/' + imageName + '.png', frame):
            print('Image saved successfully')
            input('Press enter to continue...')
            try:
                workbook = openpyxl.load_workbook('Attendance.xlsx')
                sheet = workbook['Details']
                last_row = sheet.max_row
                last_column = sheet.max_column
                sheet.cell(row=last_row + 1, column=1).value = int(roll)
                sheet.cell(row=last_row + 1, column=2).value = name
                sheet.cell(row=last_row + 1, column=3).value = email
                sheet.cell(row=last_row + 1, column=4).value = parent_name
                sheet.cell(row=last_row + 1, column=5).value = parent_email
                for sheet_name in workbook.sheetnames:
                    if sheet_name == 'Details':
                        continue
                    sheet = workbook[sheet_name]
                    last_row = sheet.max_row
                    last_column = sheet.max_column
                    sheet.cell(row=last_row + 1, column=1).value = int(roll)
                    sheet.cell(row=last_row + 1, column=2).value = name
                    for i in range(2, last_column + 1):
                        sheet.cell(row=last_row + 1, column=i).value = 'A'
                workbook.save('Attendance.xlsx')
            except:
                print(
                    'Error in saving data.\nMust be the following reasons: \n1. Attendance.xlsx is not present in the current directory\n2. Attendance.xlsx is not in the correct format\n3. Roll number is not a number'
                )
        else:
            print('Image not saved\nPlease try again')
        cv2.destroyAllWindows()
        cap.release()