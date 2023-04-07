import openpyxl 
import smtplib
from datetime import datetime
from prettytable import PrettyTable
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import os
path = "Attendance.xlsx"

def get_dict():
    wb_obj = openpyxl.load_workbook(path)
    sheet = wb_obj['Details']
    data = {}
    for row in sheet.rows:
        for values in row:
            data[values.row] = [values.value for values in row]
    return data

def get_user_row(emp_id):
    users = get_dict()
    for key, value in users.items():
        if str(value[0]) == str(emp_id):
            return key
    return -1

def getMonthNames():
    worbook = openpyxl.load_workbook(path)
    months = []
    for sheet in worbook.sheetnames:
        if sheet != 'Details':
            months.append(sheet)
    return months

def add_column():
    today = datetime.date(datetime.now())
    curr_month = today.strftime("%B")
    wb_obj = openpyxl.load_workbook(path) 
    try:
        sheet_obj = wb_obj[curr_month]
    except:
        wb_obj.create_sheet(curr_month)
        sheet_obj = wb_obj[curr_month]
        sheet_obj.cell(row=1, column=1).value = 'Roll no'
        sheet_obj.cell(row=1, column=2).value = 'Name'
        sheet_2 = wb_obj['Details']
        for i in range(2, sheet_2.max_row+1):
            sheet_obj.cell(row=i, column=1).value = sheet_2.cell(row=i, column=1).value
            sheet_obj.cell(row=i, column=2).value = sheet_2.cell(row=i, column=2).value
        wb_obj.save(path)
    last = sheet_obj.max_column
    last_column_name = get_col_name(last)
    if last_column_name != str(today):
        print("Adding new column, number of columns:", last)
        cell = sheet_obj.cell(row=1, column=last+1)
        cell.value = str(today)
        last += 1
        for i in range(2, sheet_obj.max_row+1):
            cell = sheet_obj.cell(row=i, column=last)
            cell.value = 'A'
        wb_obj.save(path)
    return last

def get_col_name(colnum):
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj[datetime.now().strftime("%B")]
    cell = sheet_obj.cell(row=1, column=colnum)
    return str(cell.value)

def write_attendance_in_xl(name, count):
    roll_no = int(name.split('_')[0])
    lastColumn = add_column()
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj[datetime.now().strftime("%B")]
    key = get_user_row(roll_no)
    if key!=-1:
        cell = sheet_obj.cell(row=key, column=lastColumn)
        print(cell.value)
        if cell.value == "A":
            cell.value = 'In-time: ' + str(datetime.now().time()).split('.')[0]
            count+=1
        elif cell.value.startswith("In-time:"):
            in_time = cell.value.split(' ')[1]
            cell.value = 'In-time: ' + str(in_time) + ' \nOut-time: ' + str(datetime.now().time()).split('.')[0]
    wb_obj.save(path)
    return count

def fetch_attendance():
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj[datetime.now().strftime("%B")]
    roll_no = int(input("Enter employee ID: "))
    key = get_user_row(roll_no)
    if key != -1:   
        attendance_records = PrettyTable(['Date', 'In-time', 'Out-time', 'Hours'])
        print("\nAttendance records for the month of ", datetime.now().strftime("%B"))
        for i in range(3, sheet_obj.max_column+1):
            cell = sheet_obj.cell(row=key, column=i)
            if cell.value.startswith("In-time:"):
                in_time = cell.value.split(' ')[1]
                out_time = cell.value.split(' ')[3]
                hours = str((datetime.strptime(out_time, '%H:%M:%S') - datetime.strptime(in_time, '%H:%M:%S')).total_seconds()/3600)[:4]
                attendance_records.add_row([get_col_name(i), in_time, out_time, hours])
            else:
                attendance_records.add_row([get_col_name(i), 'A', 'A', 0])
        print (attendance_records)
    else:
        print("\nNo record found for this roll number.")
    print("\n")
    input("Press any key to continue...")

def send_monthly_attendance(month = datetime.now().strftime("%B")):
    s = smtplib.SMTP('smtp.gmail.com', 587)
    s.starttls()
    sender_email = "d2021.piyush.chugeja@ves.ac.in"
    s.login(sender_email, os.getenv('EMAIL_PASSWORD'))
    workbook = openpyxl.load_workbook(path)
    attendance_all = workbook[month]
    member_rows = get_dict()
    for key, value in member_rows.items():
        if key != 1:
            attendance_html_table = '''
            <table border="1" style="border-collapse: collapse; width: 100%;">
                <tr>
                    <th>Date</th>
                    <th>In-time</th>
                    <th>Out-time</th>
                    <th>Hours</th>
                </tr>
            '''
            message = MIMEMultipart("alternative")
            message["From"] = "Attendance System"
            total, attended, totalHours, userHours = 0, 0, 0, 0
            for i in range(3, attendance_all.max_column+1):
                cell = attendance_all.cell(row=key, column=i)
                total += 1
                totalHours += 6
                date = attendance_all.cell(row=1, column=i).value
                if cell.value.startswith("In-time:"):
                    attended += 1
                    in_time = cell.value.split(' ')[1]
                    try:
                        out_time = cell.value.split(' ')[3]
                        hours = str((datetime.strptime(out_time, '%H:%M:%S') - datetime.strptime(in_time, '%H:%M:%S')).total_seconds()/3600)[:4]   
                        userHours += float(hours)
                        color = 'white'
                        if float(hours) < 6:
                            color = 'yellow'
                    except:
                        out_time = "Unavailable"
                        hours = "-"
                    attendance_html_table += '''
                    <tr>
                        <td>{date}</td>
                        <td>{in_time}</td>
                        <td>{out_time}</td>
                        <td style='background: {color};'>{hours}</td>
                    </tr>
                    '''.format(date=date, in_time=in_time, out_time=out_time, hours=hours, color=color)
                else:
                    attendance_html_table += '''
                    <tr>
                        <td>{date}</td>
                        <td>A</td>
                        <td>A</td>
                        <td style = 'background: red;'>0</td>
                    </tr>
                    '''.format(date=date)
            attendance = round((attended/total)*100, 2)
            requiredHours = totalHours - userHours
            if not requiredHours <= 0:
                requiredMessage = "You need to work for {requiredHours} more hours to complete your monthly target.".format(requiredHours=requiredHours)
            else:
                requiredMessage = "You have completed your monthly target."
            attendance_html_table += '''
            </table>
            <br>
            <p>
                Overall attendance for {month} is {attendance}%.
                {requiredMessage}
            '''.format(month=month, attendance=attendance, requiredMessage=requiredMessage)
            receiver = {'email':value[2], 'name':value[1]}
            
            text = """\
            Hi {receiver[name]},
            Your attendance for this month is {attendance}%.
            """.format(receiver=receiver, attendance=attendance)
            
            html = """\
            <html>
                <body>
                    <h1 align='center'>Email from attendance system</h1>
                    <hr>
                    <p>
                        Hi {receiver[name]},<br>
                        Here is your attendance record for the month of {month}:
                    </p>
                    <br>
                    {attendance_table}
                </body>
            </html>
            """.format(receiver=receiver, attendance=attendance, month=month, attendance_table=attendance_html_table)
            part1 = MIMEText(text, "plain")
            part2 = MIMEText(html, "html")
            message.attach(part1)
            message.attach(part2)
            message["To"] = receiver['email']
            message["Subject"] = "Attendance report for {month}".format(month=month)
            print("Sending email to {receiver[name]} {receiver[email]}...".format(receiver=receiver))
            s.send_message(message)
    s.quit()

def delete_user(row, emp_id):
    workbook = openpyxl.load_workbook(path)
    details = workbook['Details']
    emp_name = details.cell(row=row, column=2).value
    details.delete_rows(row)
    for sheet in workbook.sheetnames:
        if sheet != 'Details':
            workbook[sheet].delete_rows(row)
    if os.path.isfile('images/' + emp_id + '_' + emp_name.replace(' ', '_') + '.png'):
        os.remove('images/' + emp_id + '_' + emp_name.replace(' ', '_') + '.png')
    elif os.path.isfile('images/' + emp_id + '_' + emp_name.replace(' ', '_') + '.jpg'):
        os.remove('images/' + emp_id + '_' + emp_name.replace(' ', '_') + '.jpg')
    elif os.path.isfile('images/' + emp_id + '_' + emp_name.replace(' ', '_') + '.jpeg'):
        os.remove('images/' + emp_id + '_' + emp_name.replace(' ', '_') + '.jpeg')
    else:
        return False
    workbook.save(path)
    return True
    